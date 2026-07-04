"""
Rebuild International_ISA_CRM.xlsx from scratch, using the ProspectLead
database table as the single source of truth. Safe to run any time --
the database is authoritative; the xlsx is just a export/view of it for
manual review and copy-paste drafts.

Use this if the xlsx ever looks wrong, incomplete, or reverted -- it's
happened once before (2026-07-03, cause unclear) and this is the fast,
safe recovery path instead of re-researching leads from scratch.

Run with: python rebuild_crm_from_db.py
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from app import create_app
from app.models import ProspectLead

OUTPUT_PATH = Path.home() / "Downloads" / "International_ISA_CRM.xlsx"

HEADERS = ["Name", "Company/Brokerage", "Email", "Phone", "Website", "Notes",
           "Priority", "Lead Status", "Date Added", "Last Contact Date",
           "Next Action", "Draft Email", "Draft Text"]

SEGMENT_MAP = {
    "Real Estate": "Real Estate",
    "Insurance": "Insurance",
    "Architecture": "Architecture-Engineering",
    "Trades/HVAC": "Trades",
    "Trades/Plumbing": "Trades",
}

INDUSTRY_NOTES = {
    "Real Estate": "for real estate teams (lead follow-up, scheduling, admin) so agents can focus on closing instead of chasing paperwork",
    "Insurance": "for insurance agencies -- policy tracking, client follow-up, quoting support, and bilingual customer service",
    "Architecture": "for architecture and engineering firms -- client intake, scheduling, and follow-up, not design work",
    "Trades/HVAC": "for contractors -- scheduling, dispatch support, quotes, and customer follow-up so leads don't go cold while your crew's on a job site",
    "Trades/Plumbing": "for contractors -- scheduling, dispatch support, quotes, and customer follow-up so leads don't go cold while your crew's on a job site",
}


def extract_email(contact_info):
    m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", contact_info or "")
    return m.group(0) if m else ""


def extract_phone(contact_info):
    m = re.search(r"\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", contact_info or "")
    return m.group(0) if m else ""


def priority_from_score(score):
    if score is None:
        return "Medium"
    if score >= 8:
        return "High"
    if score >= 6:
        return "Medium"
    return "Low"


def draft_email(company, pain_point, industry_note):
    hook = pain_point.split(".")[0] if pain_point else "your work"
    return (
        f"Subject: Quick question about {company}\n\n"
        f"Hi there,\n\n"
        f"I came across {company} — {hook} caught my eye. "
        f"I run International ISA: bilingual, sales-trained back-office support "
        f"{industry_note}.\n\n"
        f"Would you be open to a 15-minute call this week to see if it's a fit? "
        f"No pressure either way.\n\n"
        f"[Your name]\nInternational ISA — international-isa-website.onrender.com"
    )


def draft_text(company):
    return (
        f"Hi there, this is [Your name] with International ISA — "
        f"bilingual sales & admin support. Saw {company} and thought it might "
        f"be a fit. Worth a quick call this week?"
    )


def main():
    app = create_app()
    with app.app_context():
        leads = ProspectLead.query.order_by(ProspectLead.score.desc()).all()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheets = {}
        header_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        widths = {"A": 18, "B": 32, "C": 26, "D": 16, "E": 22, "F": 46, "G": 10,
                  "H": 16, "I": 13, "J": 15, "K": 24, "L": 55, "M": 40}

        for sheet_name in ["Real Estate", "Insurance", "Architecture-Engineering", "Trades"]:
            ws = wb.create_sheet(sheet_name)
            for col, h in enumerate(HEADERS, start=1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(vertical="center", wrap_text=True)
            for col, w in widths.items():
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 30
            sheets[sheet_name] = ws

        row_counters = {name: 2 for name in sheets}

        for lead in leads:
            sheet_name = SEGMENT_MAP.get(lead.industry, "Trades")
            ws = sheets[sheet_name]
            r = row_counters[sheet_name]

            email = extract_email(lead.contact_info)
            phone = extract_phone(lead.contact_info)
            notes = f"{lead.pain_point or ''} Est. value: {lead.estimated_value or 'TBD'}. Fit: {lead.solution_fit or 'TBD'}"
            priority = priority_from_score(lead.score)
            next_action = "Send initial email" if email else "Find contact info, then send"
            industry_note = INDUSTRY_NOTES.get(lead.industry, "for your business")

            last_contact = lead.last_contact_date.strftime("%Y-%m-%d") if lead.last_contact_date else ""
            row_data = [
                "", lead.company_name, email, phone, "", notes, priority,
                lead.status, lead.date_added.strftime("%Y-%m-%d") if lead.date_added else "",
                last_contact, next_action,
                draft_email(lead.company_name, lead.pain_point, industry_note),
                draft_text(lead.company_name),
            ]
            for c, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_counters[sheet_name] += 1

        for sheet_name, ws in sheets.items():
            last_row = row_counters[sheet_name] - 1
            if last_row < 2:
                continue
            dv_status = DataValidation(
                type="list",
                formula1='"Not Contacted,Contacted,Follow-Up Needed,Interested,Not Interested,Closed-Won"',
                allow_blank=True,
            )
            ws.add_data_validation(dv_status)
            dv_status.add(f"H2:H{last_row}")
            dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
            ws.add_data_validation(dv_priority)
            dv_priority.add(f"G2:G{last_row}")

        wb.save(OUTPUT_PATH)
        for name in sheets:
            print(f"{name}: {row_counters[name] - 2} leads")
        print(f"Saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
